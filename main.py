import serial
import mysql.connector
from datetime import datetime
import os
from openpyxl import Workbook, load_workbook
import tkinter as tk
from tkinter import messagebox

# MySQL database settings
db_config = {
    'host': 'localhost',
    'user': 'YOUR USERNAME',  # Update with your username
    'password': 'YOUR PASSWORD',  # Update with your password
    'database': 'Employee'  # Use the 'Employee' database
}

class DatabaseHandler:
    def __init__(self, config):
        try:
            self.conn = mysql.connector.connect(**config)
            self.cursor = self.conn.cursor()
            print("Connected to MySQL database")
        except mysql.connector.Error as err:
            print("Error connecting to MySQL:", err)
            exit(1)

    def check_rfid_exists(self, uid):
        try:
            query = "SELECT Employee_uid, Employee_name, Employee_mobile_no FROM Registered_Employee WHERE Employee_RFID = %s"
            self.cursor.execute(query, (uid,))
            return self.cursor.fetchone()
        except mysql.connector.Error as err:
            print("Error checking RFID existence:", err)
            return None

    def insert_attendance(self, emp_name, emp_mobile, uid):
        try:
            query = "INSERT INTO Registered_Employee (Employee_name, Employee_mobile_no, Employee_RFID) VALUES (%s, %s, %s)"
            self.cursor.execute(query, (emp_name, emp_mobile, uid))
            self.conn.commit()
            # Fetch the last inserted ID
            self.cursor.execute("SELECT LAST_INSERT_ID()")
            emp_id = self.cursor.fetchone()[0]
            print("Attendance recorded for RFID:", uid)
            return emp_id
        except mysql.connector.Error as err:
            print("Error inserting data:", err)
            return None

    def close(self):
        self.cursor.close()
        self.conn.close()
        print("Database connection closed")


class ExcelHandler:
    def __init__(self, file_path):
        self.file_path = file_path
        self.ensure_file_exists()

    def ensure_file_exists(self):
        if not os.path.exists(self.file_path):
            wb = Workbook()
            ws = wb.active
            ws.append(['Employee ID', 'Employee Name', 'Mobile Number', 'RFID', 'Time'])
            wb.save(self.file_path)
            print(f"Excel file '{self.file_path}' created successfully.")

    def write_to_excel(self, data):
        wb = load_workbook(self.file_path)
        ws = wb.active
        ws.append(data)
        wb.save(self.file_path)
        print("Data written to Excel file successfully.")


class SerialHandler:
    def __init__(self, port, baud_rate):
        self.ser = serial.Serial(port, baud_rate)

    def read_line(self):
        return self.ser.readline().decode().strip()

    def write_line(self, data):
        self.ser.write(data.encode())

    def close(self):
        self.ser.close()
        print("Serial port closed")


class UserInputDialog(tk.Tk):
    def __init__(self, uid):
        super().__init__()
        self.uid = uid
        self.emp_name = None
        self.emp_mobile = None
        self.title("Register Employee")
        self.geometry("400x300")
        self.create_widgets()

    def create_widgets(self):
        tk.Label(self, text="RFID not found. Please register.").pack(pady=10)
        tk.Label(self, text=f"RFID: {self.uid}").pack(pady=5)
        
        tk.Label(self, text="Employee Name:").pack(pady=5)
        self.name_entry = tk.Entry(self)
        self.name_entry.pack(pady=5)
        
        tk.Label(self, text="Mobile Number:").pack(pady=5)
        self.mobile_entry = tk.Entry(self)
        self.mobile_entry.pack(pady=5)
        
        tk.Button(self, text="Submit", command=self.on_submit).pack(pady=20)

    def on_submit(self):
        self.emp_name = self.name_entry.get()
        self.emp_mobile = self.mobile_entry.get()
        if self.emp_name and self.emp_mobile:  # Ensure that both fields are filled
            self.destroy()
        else:
            messagebox.showwarning("Input Error", "Please fill in all fields.")

    def get_results(self):
        self.mainloop()
        return self.emp_name, self.emp_mobile


def main():
    # File path for the Excel file
    desktop_path = "C:\\Users\\NEW\\OneDrive\\Desktop\\"
    excel_file_path = os.path.join(desktop_path, 'attendance.xlsx')

    # Initialize handlers
    db_handler = DatabaseHandler(db_config)
    excel_handler = ExcelHandler(excel_file_path)
    serial_handler = SerialHandler('COM6', 9600)

    try:
        while True:
            line = serial_handler.read_line()
            if line.startswith('Card UID:'):
                uid = line.split(': ')[1]
                result = db_handler.check_rfid_exists(uid)
                if result:
                    emp_id, emp_name, emp_mobile = result
                    current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                    excel_handler.write_to_excel([emp_id, emp_name, emp_mobile, uid, current_time])
                    x="1"
                    serial_handler.write_line(x+"\n")
                    serial_handler.write_line(emp_name+"\n")
                    print("Employee Name sent to Arduino:", emp_name)
                else:
                    x="0"
                    serial_handler.write_line(x+"\n")
                    emp_name, emp_mobile = UserInputDialog(uid).get_results()
                    if emp_name and emp_mobile:
                        emp_id = db_handler.insert_attendance(emp_name, emp_mobile, uid)
                        if emp_id:
                            messagebox.showinfo("Registration Successful", f"Employee registered successfully.\nEmployee ID: {emp_id}")
                        else:
                            messagebox.showerror("Registration Error", "Failed to register employee. Please try again.")
                    else:
                        messagebox.showerror("Registration cancelled or incomplete.")
    except KeyboardInterrupt:
        print("Process interrupted by user")
    finally:
        serial_handler.close()
        db_handler.close()


if __name__ == "__main__":
    main()
